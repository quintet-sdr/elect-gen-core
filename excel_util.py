import os
import random
from openpyxl import load_workbook, Workbook
from models import Course, Student


def readCoursesInfo():
    courseFile = 'Courses.xlsx'
    courseWB = load_workbook(courseFile)
    courseSheet = courseWB[courseWB.sheetnames[0]]
    courses = []
    for i in range(2, courseSheet.max_row + 1):
        ID = courseSheet.cell(row=i, column=1).value
        name = courseSheet.cell(row=i, column=2).value
        quota = courseSheet.cell(row=i, column=3).value
        courses.append(Course(ID, name, quota))
    return courses


def readStudentsInfo(courses, students, name):
    studentFile = name
    studentWB = load_workbook(studentFile)
    studentSheet = studentWB[studentWB.sheetnames[0]]
    for i in range(2, studentSheet.max_row + 1):
        ID = studentSheet.cell(row=i, column=1).value
        name = studentSheet.cell(row=i, column=2).value
        GPA = studentSheet.cell(row=i, column=3).value
        keywords = []
        availableCourses = []
        for j in range(4, 9):
            keyword = studentSheet.cell(row=i, column=j).value
            flag = False
            for course in courses:
                if course.name.lower() == keyword.lower():
                    flag = True
                    keywords.append(course)
                    if course.ID != -1:
                        availableCourses.append(course.name)
                    break
            if not flag:
                course = Course(-1, keyword, 0)
                courses.append(course)
                keywords.append(course)
        GPA = float(GPA)
        if len(availableCourses) < 5:
            other_courses = [course.name for course in courses if course.name not in availableCourses]
            random_courses = random.sample(other_courses, 5 - len(availableCourses))
            availableCourses += random_courses

        students.append(Student(ID, name, GPA, keywords, availableCourses))
    students.sort(key=lambda student: student.GPA, reverse=True)


def writeResults(students_distributions, costs, courses):
    print("Writing results in table...")
    if "Results.xlsx" in os.listdir():
        os.remove("Results.xlsx")
    results = Workbook()
    costs_dict = {cost: students for students, cost in zip(students_distributions, costs)}
    numeration = 1
    for cost, students in costs_dict.items():
        resultsSheetResults = results.create_sheet(title=f"Result {numeration}")
        resultsSheetResults.append(
            ["Student ID", "Student Name", "Final priority", "Course Name", "Actual Course Name", "1 priority",
             "2 priority",
             "3 priority",
             "4 priority", "5 priority", "No priority", "Bad input", ""] + [course.name for course in courses])

        totalResults = [0] * 7
        totalCourseResults = [0] * len(courses)
        totalCourseQuotas = [course.quota for course in courses]
        for student in students:
            resultsSheetResults.append([student.ID, student.name, student.finalPriority, student.finalCourse,
                                        student.keywords[
                                            student.finalPriority - 1].name if student.finalPriority <= 5 else ""])
            if 1 <= student.finalPriority <= 7:
                totalResults[student.finalPriority - 1] += 1
            matching_courses = [course for course in courses if course.name == student.finalCourse]
            if matching_courses:
                totalCourseResults[courses.index(matching_courses[0])] += 1
        for j in range(0, 7):
            resultsSheetResults.cell(row=2, column=j + 6, value=totalResults[j])
        resultsSheetResults.cell(row=2, column=13, value="Distribution:")
        resultsSheetResults.cell(row=3, column=13, value="Quotas:")
        for k in range(0, len(courses)):
            resultsSheetResults.cell(row=2, column=k + 14, value=totalCourseResults[k])
        for k in range(0, len(courses)):
            resultsSheetResults.cell(row=3, column=k + 14, value=totalCourseQuotas[k])
        numeration += 1
    print("Results written in table")
    results.save("Results.xlsx")
