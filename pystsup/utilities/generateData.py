import openpyxl as xl
from openpyxl import *
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.chart import *

from .acmParser import getPath, parseFile
from pystsup.data import Student, Course


def getData(stuFile, supFile, keywordsFile="pystsup/test/acm.txt"):
    topicNames, topicPaths, topicIDs, levels = parseFile(keywordsFile)

    stuSheet = (xl.load_workbook(stuFile)).active
    supSheet = (xl.load_workbook(supFile)).active

    m = len(list(supSheet.rows)) - 1
    n = len(list(stuSheet.rows)) - 1

    numberOfStudents = n
    numberOfcourses = m

    stuRows = list(stuSheet.rows)
    del stuRows[0]
    supRows = list(supSheet.rows)
    del supRows[0]

    stuID = "student"
    supID = "course"

    students = {}
    courses = {}

    for row in stuRows:

        n -= 1
        rank = 0
        stu_List = {}

        stu = stuID + str(numberOfStudents - n)

        realID = row[0].value
        name = row[1].value
        keywords = []

        for i in range(2, len(row)):
            val = row[i].value
            val = val.lower().strip()
            if val != None:
                keywords.append(val)
            else:
                keywords.append(topicIDs[1])

        for kw in keywords:
            rank += 1
            stu_List[rank] = [kw, getPath(kw, topicNames, topicPaths, topicIDs)]

        studentObject = Student(stu, stu_List, realID, name)
        students[stu] = studentObject

    for row in supRows:

        m -= 1
        rank = 0
        sup_List = {}

        sup = supID + str(numberOfcourses - m)

        realID = row[0].value
        name = row[1].value
        quota = int(row[2].value)
        keywords = []

        for i in range(3, len(row)):
            val = row[i].value
            val = val.lower().strip()
            if val != None:
                keywords.append(val)
            else:
                keywords.append(topicIDs[1])

        for kw in keywords:
            rank += 1
            sup_List[rank] = [kw, getPath(kw, topicNames, topicPaths, topicIDs)]

        courseObject = Course(sup, sup_List, quota, realID, name)
        courses[sup] = courseObject

    return students, courses


def scanInputData(stuFile, supFile, keywordsFile):
    topicNames, topicPaths, topicIDs, levels = parseFile(keywordsFile)

    stuWB = xl.load_workbook(stuFile)
    stuSheet = (stuWB).active

    supWB = xl.load_workbook(supFile)
    supSheet = (supWB).active

    stuRows = list(stuSheet.rows)
    del stuRows[0]
    supRows = list(supSheet.rows)
    del supRows[0]

    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    nofill = PatternFill(fill_type='none')

    errorStu = 0
    errorSup = 0

    print("Scanning Student File...")
    for i in range(len(stuRows)):

        row = stuRows[i]

        for j in range(2, len(row)):
            val = row[j].value
            if (val == None):
                continue
            val = val.lower().strip()
            if val not in topicNames:

                # print(f"Found Error in Row {i+1} Keyword {j-2} = {val}")
                errorStu += 1
                stuSheet[f"{chr(65 + j)}{i + 2}"].fill = redFill
            else:
                stuSheet[f"{chr(65 + j)}{i + 2}"].fill = nofill

    print("\n\nScanning course Excel File...")

    for i in range(len(supRows)):

        row = supRows[i]

        for j in range(3, len(row)):
            val = row[j].value
            if (val == None):
                continue
            val = val.lower().strip()

            if val not in topicNames:
                # print(f"Found Error in Row {i+1} Keyword {j-2} = {val}")
                errorSup += 1
                supSheet[f"{chr(65 + j)}{i + 2}"].fill = redFill
            else:
                supSheet[f"{chr(65 + j)}{i + 2}"].fill = nofill

    print("\nScan Complete..\n==========================")
    print(f"Total Errors Found in Student Input File = {errorStu}")
    print(f"Total Errors Found in course Input File = {errorSup}")

    stuWB.save(stuFile)
    supWB.save(supFile)

    if (errorStu > 0 or errorSup > 0):
        print("\nErrors marked and saved in provided excel files\n")
    return errorStu, errorSup


def createExcelFile(filename, data, student):
    wb = Workbook()
    ws = wb.active
    if student:
        heading = ("Student ID", "Student Name", "Keyword 1", "Keyword 2", "Keyword 3", "Keyword 4", "Keyword 5")

    else:
        heading = ("course ID", "course Name", "Quota", "Keyword 1", "Keyword 2", "Keyword 3", "Keyword 4", "Keyword 5")

    ws.append(heading)

    for i in data:
        ws.append(i)

    wb.save(filename)


def writeFrontier(filename, front, metricData, courses, students):
    filename += '.xlsx'

    wb = Workbook()

    # Creating the Statistics Sheet

    ws = wb.active
    ws.title = "GA Statistics"
    ws.append(("Total Number of Generations", metricData['numberOfGenerations']))
    ws.append(("Total Time Taken (s)", metricData['total_time_generations']))
    ws.append(("Time Per Generation (s)", metricData['avg_time_generation']))
    ws.append((None, None))
    ws.append(("Initial Maximum Student Fitness (Fst)", metricData['initial_maxFst']))
    ws.append(("Final Maximum Student Fitness (Fst)", metricData['final_maxFst']))
    ws.append((None, None))
    ws.append(("Initial Minimum Student Fitness (Fst)", metricData['initial_minFst']))
    ws.append(("Final Minimum Student Fitness (Fst)", metricData['final_minFst']))
    ws.append((None, None))
    ws.append(("Initial Maximum course Fitness (Fsup)", metricData['initial_maxFsup']))
    ws.append(("Final Maximum course Fitness (Fsup)", metricData['final_maxFsup']))
    ws.append((None, None))
    ws.append(("Initial Minimum course Fitness (Fsup)", metricData['initial_minFsup']))
    ws.append(("Final Minimum course Fitness (Fsup)", metricData['final_minFsup']))
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 17

    wb.create_sheet("Overview")
    ws = wb["Overview"]

    heading = ("Sheet Number", "Student Fitness", "course Fitness")
    ws.append(heading)

    maxFst = max(front, key=lambda x: x.getFst()).getFst()
    maxFsup = max(front, key=lambda x: x.getFsup()).getFsup()
    minFst = min(front, key=lambda x: x.getFst()).getFst()
    minFsup = min(front, key=lambda x: x.getFsup()).getFsup()

    front = set(front)
    n = len(front)

    if len(front) > 1:
        front = sorted(front, key=lambda x: (
                    ((x.getFsup() - minFsup) / (maxFsup - minFsup)) * ((x.getFst() - minFst) / (maxFst - minFst))),
                       reverse=True)

    count = 1

    sheetnames = []
    widths = [25, 25, 15, 15]
    letters = ['A', 'B', 'C', 'D']

    for sol in front:

        sheetName = "Solution " + str(count)
        wb.create_sheet(sheetName)
        sheetnames.append((sheetName, sol.getFst(), sol.getFsup()))
        ws = wb[sheetName]

        supEdges = sol.getGraph().getEdges()

        heading = ("course Name", "Student Name", "course ID", "Student ID")

        ws.append(heading)

        for sup in supEdges:

            supName = courses[sup].getcourseName()
            supID = courses[sup].getRealID()

            for stu in supEdges[sup]:
                stuName = students[stu].getStudentName()
                stuID = students[stu].getRealID()

                tempRow = (supName, stuName, supID, stuID)

                ws.append(tempRow)

        for i, width in enumerate(widths):
            ws.column_dimensions[letters[i]].width = width

        count += 1

    ws = wb['Overview']

    for i in sheetnames:
        ws.append(i)

    ws["A" + str(n + 4)] = "Diversity Metric"
    ws["B" + str(n + 4)] = metricData['diversity']

    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17

    wb.save(filename)
