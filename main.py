import copy
import os
import numpy as np
from openpyxl import load_workbook
import random
from ga import genetic_algorithm
import math
from openpyxl import Workbook
from multiprocessing import Pool


def calculate_success_rate(num_students):
    if 0 <= num_students <= 16:
        return 0
    elif 17 <= num_students <= 20:
        return math.exp(num_students - 16)
    elif 21 <= num_students <= 28:
        return math.exp(4 - (num_students - 20))
    elif 29 <= num_students <= 34:
        return math.exp(num_students - 28)
    elif 35 <= num_students <= 40:
        return math.exp(6 - (num_students - 34))
    elif 41 <= num_students <= 56:
        return math.exp(num_students - 40)
    elif 57 <= num_students <= 72:
        return math.exp(16 - (num_students - 56))
    elif 73 <= num_students <= 88:
        return math.exp(num_students - 72)
    elif 89 <= num_students <= 104:
        return math.exp(16 - (num_students - 88))
    elif 105 <= num_students <= 120:
        return math.exp(num_students - 104)
    else:
        return 0


success_rate_dict = {i: calculate_success_rate(i) for i in range(0, 1000)}


class Student:
    def __init__(self, ID, name, GPA, keywords, availableCourses):
        self.isDistributed = False
        self.ID = ID
        self.name = name
        self.GPA = GPA
        self.keywords = keywords
        self.finalCourse = ""
        self.finalPriority = 6
        self.availableCourses = availableCourses


class Course:
    def __init__(self, ID, name, quota):
        self.ID = ID
        self.name = name
        self.quota = quota
        self.students = []


def memoize(func):
    cache = dict()

    def memoized_func(*args):
        hashable_args = tuple(tuple(x) if isinstance(x, list) else x for x in args)
        if hashable_args in cache:
            return cache[hashable_args]
        result = func(*args)
        cache[hashable_args] = result
        return result

    return memoized_func


def readCoursesInfo():
    courseFile = 'Courses table.xlsx'
    courseWB = load_workbook(courseFile)
    courseSheet = courseWB[courseWB.sheetnames[0]]
    courses = []
    for i in range(2, courseSheet.max_row + 1):
        ID = courseSheet.cell(row=i, column=1).value
        name = courseSheet.cell(row=i, column=2).value
        quota = courseSheet.cell(row=i, column=3).value
        courses.append(Course(ID, name, quota))
    return courses


def readStudentsInfo(courses, students, name):
    studentFile = name
    studentWB = load_workbook(studentFile)
    studentSheet = studentWB[studentWB.sheetnames[0]]
    for i in range(2, studentSheet.max_row + 1):
        ID = studentSheet.cell(row=i, column=1).value
        name = studentSheet.cell(row=i, column=2).value
        GPA = studentSheet.cell(row=i, column=3).value
        keywords = []
        availableCourses = []
        for j in range(4, 9):
            keyword = studentSheet.cell(row=i, column=j).value
            flag = False
            for course in courses:
                if course.name.lower() == keyword.lower():
                    flag = True
                    keywords.append(course)
                    if course.ID != -1:
                        availableCourses.append(course.name)
                    break
            if not flag:
                course = Course(-1, keyword, 0)
                courses.append(course)
                keywords.append(course)
        GPA = float(GPA)
        if len(availableCourses) < 5:
            other_courses = [course.name for course in courses if course.name not in availableCourses]
            random_courses = random.sample(other_courses, 5 - len(availableCourses))
            availableCourses += random_courses

        students.append(Student(ID, name, GPA, keywords, availableCourses))
    students.sort(key=lambda student: student.GPA, reverse=True)


def writeResults(students_distributions, costs, courses):
    if "Results.xlsx" in os.listdir():
        os.remove("Results.xlsx")
    results = Workbook()
    costs_dict = {cost: students for students, cost in zip(students_distributions, costs)}

    for cost, students in costs_dict.items():
        resultsSheetResults = results.create_sheet(title=f"Result {cost}")
        resultsSheetResults.append(
            ["Student ID", "Student Name", "Final priority", "Course Name", "Actual Course Name", "1 priority",
             "2 priority",
             "3 priority",
             "4 priority", "5 priority", "No priority", "Bad input", ""] + [course.name for course in courses])

        totalResults = [0] * 7
        totalCourseResults = [0] * len(courses)
        totalCourseQuotas = [course.quota for course in courses]
        for student in students:
            resultsSheetResults.append([student.ID, student.name, student.finalPriority, student.finalCourse,
                                        student.keywords[student.finalPriority - 1].name])
            if 1 <= student.finalPriority <= 7:
                totalResults[student.finalPriority - 1] += 1
            totalCourseResults[
                courses.index([course for course in courses if course.name == student.finalCourse][0])] += 1
        for j in range(0, 7):
            resultsSheetResults.cell(row=2, column=j + 6, value=totalResults[j])
        resultsSheetResults.cell(row=2, column=13, value="Distribution:")
        resultsSheetResults.cell(row=3, column=13, value="Quotas:")
        for k in range(0, len(courses)):
            resultsSheetResults.cell(row=2, column=k + 14, value=totalCourseResults[k])
        for k in range(0, len(courses)):
            resultsSheetResults.cell(row=3, column=k + 14, value=totalCourseQuotas[k])

    results.save("Results.xlsx")


@memoize
def costFunction(students, courses):
    cost = 0
    student_priorities = np.array([student.finalPriority for student in students])
    student_GPAs = np.array([max(student.GPA, 0.1) for student in students])
    cost += np.sum((student_priorities ** 2) / student_GPAs)

    for course in courses:
        numOfStudents = len(course.students)
        cost += success_rate_dict[numOfStudents] ** 4
    return cost


def get_student_by_id(students):
    return {s.ID: s for s in students}


def calculate_fitness_helper(args, students, courses):
    individual = args
    fitness = costFunction(students, courses)
    return fitness


def calculate_fitness_parallel(individuals, students, courses):
    with Pool() as pool:
        fitnesses = pool.starmap(calculate_fitness_helper,
                                 [(individual, students, courses) for individual in individuals])
    return fitnesses


def Distribute(students, errorCourse, courses):
    random.shuffle(students)
    students.sort(key=lambda student: student.GPA, reverse=True)
    for student in students:
        if not student.isDistributed:
            for courseName in student.availableCourses:
                for course in courses:
                    if courseName.lower() == course.name.lower() and course.quota > 0:
                        course.quota -= 1
                        course.students.append(student)
                        student.isDistributed = True
                        student.finalCourse = course.name
                        student.finalPriority = student.availableCourses.index(courseName) + 1
                        break
                if student.isDistributed:
                    break
    for student in students:
        if not student.isDistributed:
            other_courses = [course for course in courses if course.name not in student.availableCourses]
            other_courses.sort(key=lambda course: len(course.students))
            if other_courses:
                course = other_courses[0]
                course.students.append(student)
                student.isDistributed = True
                student.finalCourse = course.name
                student.finalPriority = len(student.availableCourses)
    cost = costFunction(students, courses)
    return students, cost


def improveDistribution(students, courses, errorCourse):
    noImprovements = 0
    cost = costFunction(students, courses)
    print("Initial cost: ", cost)
    while noImprovements < 1:
        newStudents = [copy.copy(s) for s in students]
        newCourses = [copy.copy(c) for c in courses]
        sortedStudents = sorted(students, key=lambda s: s.finalPriority, reverse=True)
        numStudents = max(int(len(students) * cost / 1000), 1)
        selectedStudents = sortedStudents[:numStudents]
        for student in selectedStudents:
            if student.finalPriority != 1 and student.finalPriority != 7:
                for findStudent in newStudents:
                    if findStudent.name == student.name:
                        findStudent.finalPriority = min(findStudent.finalPriority - 1,
                                                        len(findStudent.availableCourses) - 1)
                        if findStudent.finalPriority < len(findStudent.availableCourses):
                            findStudent.finalCourse = findStudent.availableCourses[findStudent.finalPriority]
                            findStudent.isDistributed = True
                            for course in newCourses:
                                if findStudent.availableCourses[findStudent.finalPriority] == course.name:
                                    course.quota -= 1
                                    course.students.append(student)
        Distribute(newStudents, errorCourse, newCourses)
        newCost = costFunction(newStudents, newCourses)

        if newCost < cost:
            print("Cost improved: ", newCost)
            cost = newCost
            students = newStudents
            courses = newCourses
            noImprovements = 0
        else:
            noImprovements += 1
            if noImprovements % 100 == 0:
                print("Without improvements: ", noImprovements)
    return students, cost


def selectDistribution(cmd):
    best_distribution_students = []
    best_distribution_cost = None
    courses = readCoursesInfo()
    students = []
    readStudentsInfo(courses, students, "Students table.xlsx")
    readStudentsInfo(courses, students, "Students table 2.xlsx")
    students_dict = get_student_by_id(students)
    if cmd == 1:
        best_distribution = genetic_algorithm(students, courses)
        best_distribution_students = [students_dict[student_id] for student_id, _ in best_distribution]
        best_distribution_cost = costFunction(best_distribution_students, courses)
    elif cmd == 2:
        best_distribution_students, best_distribution_cost = startBasicAlgorithm(students, courses)
    return best_distribution_students, best_distribution_cost


def startBasicAlgorithm(students, courses):
    clearCourses = courses.copy()
    clearStudents = students.copy()
    errorCourse = Course(-1, "ERROR", 0)
    students_distributions = []
    costs = []
    for _ in range(5):
        students_copy = [copy.deepcopy(s) for s in students]
        courses_copy = [copy.deepcopy(c) for c in courses]
        students_copy, cost = Distribute(students_copy, errorCourse, courses_copy)
        students_copy, cost = improveDistribution(students_copy, courses_copy, clearStudents)
        students_distributions.append(students_copy)
        costs.append(cost)
    sorted_indices = np.argsort(costs)
    students_distributions = [students_distributions[i] for i in sorted_indices]
    costs = [costs[i] for i in sorted_indices]
    return students_distributions[:10], costs[:10]
